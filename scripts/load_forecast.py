"""
Carga el forecast semanal de horas desde el export
"Horas y % Utilización por Recurso y Proyecto".

Uso:
    pip install openpyxl psycopg[binary] python-dotenv
    python scripts/load_forecast.py path/al/archivo.xlsx

Estructura esperada del Excel:
  - Columna E: Empleado  (formato "Apellido, Nombre GUID")
  - Columna F: Proyecto  (formato "Nombre Proyecto – engagement_id")
  - Columna G en adelante: grupos de 3 columnas por semana:
      col+0: Horas efectivas
      col+1: Horas facturables
      col+2: % Utilización  (se descarta)
  La fila con las fechas de inicio de semana se detecta automáticamente.

Duplicados:
  - Si el mismo (empleado, engagement, semana) aparece varias veces dentro
    del mismo Excel, se mantiene la última ocurrencia.
  - Si se vuelve a cargar un Excel que solapa semanas ya existentes en DB,
    los valores se sobreescriben (ON CONFLICT DO UPDATE).

Requiere DATABASE_URL en .env.local apuntando a Postgres/Supabase.
"""
from __future__ import annotations

import os
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import psycopg
from psycopg.rows import dict_row

try:
    from dotenv import load_dotenv
    load_dotenv(".env.local")
    load_dotenv(".env")
except ImportError:
    pass

# Columnas fijas (1-indexed, igual que openpyxl)
EMPLOYEE_COL  = 5   # E
PROJECT_COL   = 6   # F
DATA_START_COL = 7  # G  (primer col de datos semanales)


# ── Helpers ──────────────────────────────────────────────────────────────────

def s(v: Any) -> str | None:
    if v is None or v == "":
        return None
    r = str(v).strip()
    return r or None


def num(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def to_date(v: Any) -> date | None:
    """Convierte cualquier representación de fecha en un objeto date."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        return datetime.fromisoformat(str(v)).date()
    except ValueError:
        pass
    # Intenta formatos comunes
    for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(str(v).strip(), fmt).date()
        except ValueError:
            continue
    return None


def parse_employee(cell: str | None) -> tuple[str | None, str | None]:
    """
    'SMITH, JOHN ABC123456'  → ('SMITH, JOHN', 'ABC123456')
    'SMITH, JOHN - ABC123456' → ('SMITH, JOHN', 'ABC123456')
    El GUID es el último token alfanumérico separado por espacio o guion.
    """
    if not cell:
        return None, None
    cell = cell.strip()
    # Separador puede ser espacio, guion o guion largo
    m = re.match(r'^(.*?)[\s\-–]+([A-Za-z0-9]{4,})\s*$', cell)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return cell, None


def parse_engagement(cell: str | None) -> tuple[str | None, str | None]:
    """
    'Nombre Proyecto – ENG123456' → ('Nombre Proyecto', 'ENG123456')
    'Nombre Proyecto - ENG123456' → ('Nombre Proyecto', 'ENG123456')
    'Nombre Proyecto ENG123456'   → ('Nombre Proyecto', 'ENG123456')
    """
    if not cell:
        return None, None
    cell = cell.strip()
    m = re.match(r'^(.*?)[\s\-–]+([A-Za-z0-9]{4,})\s*$', cell)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return cell, None


# ── Parser principal ──────────────────────────────────────────────────────────

def load(xlsx_path: str, dsn: str) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Usar la primera hoja por defecto (el fichero de forecast suele tener una)
    ws = wb.active
    print(f"Hoja activa: '{ws.title}' ({ws.max_row} filas × {ws.max_column} cols)")

    # ── Detectar fila de cabeceras de semana ──────────────────────────────────
    # Buscamos la primera fila donde las columnas G, G+3, G+6... tienen fechas.
    header_row_idx: int | None = None
    week_cols: list[tuple[int, date]] = []  # [(col_1indexed, date), ...]

    for row_idx in range(1, min(ws.max_row + 1, 30)):
        potential: list[tuple[int, date]] = []
        col = DATA_START_COL
        while col <= ws.max_column:
            d = to_date(ws.cell(row_idx, col).value)
            if d:
                potential.append((col, d))
            col += 3
        if len(potential) >= 2:
            header_row_idx = row_idx
            week_cols = potential
            break

    if not week_cols:
        print("ERROR: No se encontró ninguna fila con fechas de semana en las columnas G, J, M…")
        print("Revisa que el Excel tenga el formato esperado.")
        sys.exit(4)

    print(f"Fila de cabeceras detectada: {header_row_idx}")
    print(f"Semanas encontradas: {len(week_cols)}  "
          f"({week_cols[0][1]} → {week_cols[-1][1]})")

    # ── Parsear filas de datos ────────────────────────────────────────────────
    employees:  dict[str, str | None] = {}   # gui → name
    engagements: dict[str, str | None] = {}  # eng_id → name

    # Usamos dict con clave (gui, eng_id, semana) para deduplicar intra-Excel
    dedup: dict[tuple[str, str, date], dict] = {}

    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        emp_cell = s(ws.cell(row_idx, EMPLOYEE_COL).value)
        prj_cell = s(ws.cell(row_idx, PROJECT_COL).value)

        # Saltar filas vacías o totales
        if not emp_cell and not prj_cell:
            continue

        emp_name, emp_gui = parse_employee(emp_cell)
        eng_name, eng_id  = parse_engagement(prj_cell)

        if not emp_gui or not eng_id:
            # Fila sin GUID de empleado o sin ID de engagement → saltar
            continue

        employees[emp_gui]   = emp_name
        engagements[eng_id]  = eng_name

        for col, wk_date in week_cols:
            eff  = num(ws.cell(row_idx, col).value)
            bill = num(ws.cell(row_idx, col + 1).value)
            if eff is None and bill is None:
                continue
            key = (emp_gui, eng_id, wk_date)
            dedup[key] = {
                "employee_gui":    emp_gui,
                "engagement_id":   eng_id,
                "week_start_date": wk_date,
                "effective_hours": eff,
                "billable_hours":  bill,
            }

    forecast_rows = list(dedup.values())
    intra_dupes   = sum(1 for _ in range(len(forecast_rows))) - len(dedup)  # always 0 here; dedup is already applied

    print(f"Empleados únicos:    {len(employees)}")
    print(f"Engagements únicos:  {len(engagements)}")
    print(f"Filas de forecast:   {len(forecast_rows)}")

    if not forecast_rows:
        print("No hay filas de datos. Verifica el formato del Excel.")
        sys.exit(0)

    # ── Cargar en base de datos ───────────────────────────────────────────────
    with psycopg.connect(dsn, row_factory=dict_row, autocommit=False) as conn, \
         conn.cursor() as cur:

        cur.execute("set search_path = te, public;")

        # Stub para engagements/proyectos desconocidos
        cur.execute(
            "insert into dim_client (client_id, client_name) "
            "values ('_FORECAST_', 'Forecast Only') on conflict do nothing;"
        )
        cur.execute(
            "insert into dim_opportunity (opportunity_id, opportunity_name, client_id) "
            "values ('_FORECAST_', 'Forecast Only', '_FORECAST_') on conflict do nothing;"
        )
        cur.execute(
            "insert into dim_project (project_id, project_name, opportunity_id) "
            "values ('_FORECAST_', 'Forecast Only', '_FORECAST_') on conflict do nothing;"
        )

        # Upsert empleados (sin sobreescribir datos ya enriquecidos de T&E)
        cur.executemany(
            """
            insert into dim_employee (employee_gui, employee_name)
            values (%s, %s)
            on conflict (employee_gui) do update
              set employee_name = coalesce(excluded.employee_name, dim_employee.employee_name);
            """,
            [(k, v) for k, v in employees.items()],
        )

        # Upsert engagements (usa stub si no existe jerarquía completa)
        cur.executemany(
            """
            insert into dim_engagement (engagement_id, engagement_name, project_id)
            values (%s, %s, '_FORECAST_')
            on conflict (engagement_id) do update
              set engagement_name = coalesce(excluded.engagement_name, dim_engagement.engagement_name);
            """,
            [(k, v) for k, v in engagements.items()],
        )

        # Upsert filas de forecast
        cur.executemany(
            """
            insert into fact_forecast_week
              (employee_gui, engagement_id, week_start_date, effective_hours, billable_hours)
            values (%s, %s, %s, %s, %s)
            on conflict (employee_gui, engagement_id, week_start_date) do update
              set effective_hours = excluded.effective_hours,
                  billable_hours  = excluded.billable_hours,
                  loaded_at       = now();
            """,
            [
                (
                    r["employee_gui"],
                    r["engagement_id"],
                    r["week_start_date"],
                    r["effective_hours"],
                    r["billable_hours"],
                )
                for r in forecast_rows
            ],
        )

        conn.commit()
        print(f"OK: {len(forecast_rows)} filas de forecast cargadas correctamente.")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("uso: python scripts/load_forecast.py <ruta.xlsx>")
        sys.exit(1)

    dsn = os.environ.get("DATABASE_URL") or os.environ.get("SUPABASE_DB_URL")
    if not dsn:
        print("ERROR: define DATABASE_URL (o SUPABASE_DB_URL) en .env.local")
        sys.exit(2)

    path = Path(sys.argv[1]).expanduser().resolve()
    if not path.exists():
        print(f"ERROR: no existe el fichero {path}")
        sys.exit(3)

    load(str(path), dsn)
