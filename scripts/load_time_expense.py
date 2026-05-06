"""
Carga el libro "Detail" del export Time & Expense en Supabase / Postgres.

Uso:
    pip install openpyxl psycopg[binary] python-dotenv
    python scripts/load_time_expense.py path/al/archivo.xlsx

Requiere variable de entorno DATABASE_URL apuntando a Postgres (Supabase).
Lee .env.local si existe.
"""
from __future__ import annotations

import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
import psycopg
from psycopg.rows import dict_row

try:
    from dotenv import load_dotenv
    load_dotenv(".env.local")
    load_dotenv(".env")
except ImportError:
    pass

HEADER_ROW = 8
SHEET = "Detail"


# ---------- helpers ----------------------------------------------------------

def split_name_id(value: str | None) -> tuple[str | None, str | None]:
    """'WINNING RESULTS S.A. - 400107939' -> ('WINNING RESULTS S.A.', '400107939')."""
    if not value:
        return None, None
    m = re.match(r"^(.*?)[\s\-]+([A-Za-z0-9]+)\s*$", str(value).strip())
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return str(value).strip(), None


def to_date(v) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return datetime.fromisoformat(str(v)).date()


def num(v):
    if v in (None, ""):
        return None
    return float(v)


def s(v):
    if v is None:
        return None
    v = str(v).strip()
    return v or None


# ---------- main -------------------------------------------------------------

def load(xlsx_path: str, dsn: str) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET]
    headers = [ws.cell(HEADER_ROW, c).value for c in range(1, ws.max_column + 1)]

    rows: list[dict] = []
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        rec = {headers[i]: ws.cell(r, i + 1).value for i in range(len(headers))}
        # Saltar subtotales / vacíos
        if not rec.get("Engagement ID"):
            continue
        if not rec.get("Transaction Type") and not rec.get("Charged Hours / Tech Quantity") \
                and not rec.get("Expense Amount"):
            continue
        rows.append(rec)

    print(f"Filas a cargar: {len(rows)}")

    with psycopg.connect(dsn, row_factory=dict_row, autocommit=False) as conn, conn.cursor() as cur:
        cur.execute("set search_path = te, public;")

        # Diccionarios para deduplicar dimensiones en memoria
        clients, opps, projects, engagements = {}, {}, {}, {}
        employees, vendors, accounts, activities, categories = {}, {}, {}, {}, {}

        for r in rows:
            cid = s(r.get("Client ID"))
            if cid:
                clients[cid] = s(r.get("Client Name"))

            oid = s(r.get("Opportunity ID"))
            if oid:
                opps[oid] = (s(r.get("Opportunity Name")), cid)

            pid = s(r.get("Project ID"))
            if pid:
                projects[pid] = (s(r.get("Project Name")), oid)

            eid = s(r.get("Engagement ID"))
            if eid:
                engagements[eid] = (
                    s(r.get("Engagement Name")), pid,
                    s(r.get("Service Line")), s(r.get("Country / Region")),
                )

            gui = s(r.get("Employee GUI / Tech Product ID"))
            if gui:
                employees[gui] = (
                    s(r.get("Employee / Tech Product Name")),
                    s(r.get("GDS")),
                    s(r.get("Cost Center ")),
                    s(r.get("Employee Region")),
                    s(r.get("Employee Business Unit")),
                    s(r.get("Rank / Method")) or None,
                    s(r.get("Grade")) or None,
                )

            vname, vid = split_name_id(r.get("Vendor Name / ID"))
            if vid:
                vendors[vid] = vname

            aname, aid = split_name_id(r.get("Account Name / ID"))
            if aid:
                accounts[aid] = aname

            ac = s(r.get("Activity Code"))
            if ac:
                activities[ac] = s(r.get("Activity Code Description"))

            cc = s(r.get("Category Code"))
            if cc:
                categories[cc] = (
                    s(r.get("Category Code Description")),
                    s(r.get("Sub Category Description")),
                )

        # ---- upsert dimensiones --------------------------------------------
        cur.executemany(
            "insert into dim_client (client_id, client_name) values (%s,%s) "
            "on conflict (client_id) do update set client_name = excluded.client_name;",
            [(k, v) for k, v in clients.items()],
        )
        cur.executemany(
            "insert into dim_opportunity (opportunity_id, opportunity_name, client_id) "
            "values (%s,%s,%s) on conflict (opportunity_id) do update "
            "set opportunity_name = excluded.opportunity_name, client_id = excluded.client_id;",
            [(k, v[0], v[1]) for k, v in opps.items()],
        )
        cur.executemany(
            "insert into dim_project (project_id, project_name, opportunity_id) "
            "values (%s,%s,%s) on conflict (project_id) do update "
            "set project_name = excluded.project_name, opportunity_id = excluded.opportunity_id;",
            [(k, v[0], v[1]) for k, v in projects.items()],
        )
        cur.executemany(
            "insert into dim_engagement (engagement_id, engagement_name, project_id, service_line, country_region) "
            "values (%s,%s,%s,%s,%s) on conflict (engagement_id) do update set "
            "engagement_name = excluded.engagement_name, project_id = excluded.project_id, "
            "service_line = excluded.service_line, country_region = excluded.country_region;",
            [(k, *v) for k, v in engagements.items()],
        )

        # Asegurar ranks/grades nuevos
        cur.executemany(
            "insert into dim_rank (rank_code) values (%s) on conflict do nothing;",
            [(v[5],) for v in employees.values() if v[5]],
        )
        cur.executemany(
            "insert into dim_grade (grade) values (%s) on conflict do nothing;",
            [(v[6],) for v in employees.values() if v[6]],
        )

        cur.executemany(
            "insert into dim_employee (employee_gui, employee_name, gds, cost_center, "
            "employee_region, business_unit, rank_code, grade) "
            "values (%s,%s,%s,%s,%s,%s,%s,%s) on conflict (employee_gui) do update set "
            "employee_name = excluded.employee_name, gds = excluded.gds, cost_center = excluded.cost_center, "
            "employee_region = excluded.employee_region, business_unit = excluded.business_unit, "
            "rank_code = excluded.rank_code, grade = excluded.grade;",
            [(k, *v) for k, v in employees.items()],
        )
        cur.executemany(
            "insert into dim_vendor (vendor_id, vendor_name) values (%s,%s) "
            "on conflict (vendor_id) do update set vendor_name = excluded.vendor_name;",
            list(vendors.items()),
        )
        cur.executemany(
            "insert into dim_account (account_id, account_name) values (%s,%s) "
            "on conflict (account_id) do update set account_name = excluded.account_name;",
            list(accounts.items()),
        )
        cur.executemany(
            "insert into dim_activity (activity_code, activity_description) values (%s,%s) "
            "on conflict (activity_code) do update set activity_description = excluded.activity_description;",
            list(activities.items()),
        )
        cur.executemany(
            "insert into dim_category (category_code, category_description, sub_category_description) "
            "values (%s,%s,%s) on conflict (category_code) do update set "
            "category_description = excluded.category_description, "
            "sub_category_description = excluded.sub_category_description;",
            [(k, v[0], v[1]) for k, v in categories.items()],
        )

        # ---- hechos --------------------------------------------------------
        time_rows, expense_rows = [], []
        for r in rows:
            tt = s(r.get("Transaction Type"))
            hours = num(r.get("Charged Hours / Tech Quantity"))
            amount = num(r.get("Expense Amount"))

            eng = s(r.get("Engagement ID"))
            gui = s(r.get("Employee GUI / Tech Product ID"))
            tx_date = to_date(r.get("Transaction Date"))
            ac_date = to_date(r.get("Accounting Date"))
            wk_date = to_date(r.get("Week Ending Date"))
            activity = s(r.get("Activity Code"))

            if tt == "Labor" or hours:
                time_rows.append((
                    eng, gui,
                    s(r.get("Rank / Method")) or None,
                    s(r.get("Grade")) or None,
                    tx_date, ac_date, wk_date,
                    hours,
                    num(r.get("NSR / Tech Revenue")),
                    num(r.get("EAF Reserve Allocation")),
                    num(r.get("ANSR / Tech Revenue")),
                    num(r.get("Labor Cost")),
                    num(r.get("Labor Cost Rate")),
                    num(r.get("Tech Uplift Cost")),
                    num(r.get("Tech Product Cost")),
                    num(r.get("Tech Product Cost Rate")),
                    num(r.get("Margin Cost")),
                    num(r.get("Margin Cost Rate")),
                    num(r.get("Rate Card Rate")),
                    num(r.get("Rate Card Amount")),
                    activity,
                    "Labor",
                    bool(r.get("Relieved Flag")),
                ))
            elif amount is not None:
                _, vid = split_name_id(r.get("Vendor Name / ID"))
                _, aid = split_name_id(r.get("Account Name / ID"))
                expense_rows.append((
                    eng, vid, aid, tt,
                    gui if gui and gui != "000000000" else None,
                    tx_date, ac_date, wk_date,
                    amount,
                    s(r.get("Expense Description")),
                    s(r.get("From / Origin ")),
                    s(r.get("To / Destination")),
                    s(r.get("Trip ID")),
                    s(r.get("Journal ID")),
                    s(r.get("Voucher ID")),
                    activity,
                    s(r.get("Category Code")) or None,
                ))

        cur.execute("truncate fact_time_charge restart identity;")
        cur.execute("truncate fact_expense restart identity;")

        cur.executemany(
            """
            insert into fact_time_charge (
              engagement_id, employee_gui, rank_code, grade,
              transaction_date, accounting_date, week_ending_date,
              charged_hours, nsr_revenue, eaf_reserve_allocation, ansr_revenue,
              labor_cost, labor_cost_rate, tech_uplift_cost,
              tech_product_cost, tech_product_cost_rate,
              margin_cost, margin_cost_rate,
              rate_card_rate, rate_card_amount,
              activity_code, transaction_type_code, relieved_flag
            ) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            on conflict on constraint uq_time_charge do nothing;
            """,
            time_rows,
        )

        cur.executemany(
            """
            insert into fact_expense (
              engagement_id, vendor_id, account_id, transaction_type_code,
              employee_gui, transaction_date, accounting_date, week_ending_date,
              expense_amount, expense_description, origin, destination,
              trip_id, journal_id, voucher_id, activity_code, category_code
            ) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);
            """,
            expense_rows,
        )

        conn.commit()
        print(f"OK: {len(time_rows)} imputaciones, {len(expense_rows)} gastos cargados.")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("uso: python scripts/load_time_expense.py <ruta.xlsx>")
        sys.exit(1)

    dsn = os.environ.get("DATABASE_URL") or os.environ.get("SUPABASE_DB_URL")
    if not dsn:
        print("ERROR: define DATABASE_URL en .env.local (cadena Postgres de Supabase)")
        sys.exit(2)

    path = Path(sys.argv[1]).expanduser().resolve()
    if not path.exists():
        print(f"ERROR: no existe {path}")
        sys.exit(3)

    load(str(path), dsn)
